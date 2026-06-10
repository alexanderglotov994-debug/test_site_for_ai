#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import csv
import sys
import openpyxl

def main():
    # Настройка кодировки вывода для корректного отображения кириллицы в консоли Windows
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

    excel_path = 'Книга1.xlsx'
    csv_path = os.path.join('data', 'cases.csv')

    print(f"Запуск экспорта кейсов из {excel_path} в {csv_path}...")

    if not os.path.exists(excel_path):
        print(f"Ошибка: Файл {excel_path} не найден в текущей директории.")
        sys.exit(1)

    try:
        # Загрузка книги Excel
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Выбор листа
        sheet_name = 'Лист1'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            print(f"Лист '{sheet_name}' не найден, используем активный лист '{ws.title}'.")

        # Чтение всех строк
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print("Ошибка: Таблица Excel пуста.")
            sys.exit(1)

        # Чтение заголовков
        header = rows[0]
        header = [str(h).strip() if h is not None else "" for h in header]
        print("Обнаруженные заголовки в Excel:", header)

        # Ожидаемое сопоставление колонок
        expected_columns = {
            'title': 'Название',
            'url': 'Ссылка',
            'description': 'Текст',
            'industries': 'Фильтры отрасль',
            'products': 'Фильтры продукт',
            'divisions': 'Фильтры подразделение'
        }

        indices = {}
        for key, name in expected_columns.items():
            if name in header:
                indices[key] = header.index(name)
            else:
                # Попробуем нестрогое совпадение, если точное не найдено
                found = False
                for idx, h in enumerate(header):
                    if name.lower() in h.lower():
                        indices[key] = idx
                        print(f"Предупреждение: Колонка '{name}' не найдена точно, используем '{h}'")
                        found = True
                        break
                if not found:
                    print(f"Ошибка: Обязательная колонка '{name}' отсутствует в Excel файле.")
                    sys.exit(1)

        cases = []
        for r_idx, row in enumerate(rows[1:], start=2):
            # Пропуск полностью пустых строк
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            def get_cell(key):
                idx = indices[key]
                if idx < len(row):
                    val = row[idx]
                    return str(val).strip() if val is not None else ""
                return ""

            title = get_cell('title')
            url = get_cell('url')
            description = get_cell('description')

            # Обработка фильтров: разделение по запятым, очистка пробелов и точек на конце, объединение через |
            def process_filter(field_name):
                raw = get_cell(field_name)
                if not raw or raw.lower() == 'none':
                    return ""
                
                items = []
                # Разделяем по запятым или точкам с запятой
                for item in raw.replace(';', ',').split(','):
                    item_clean = " ".join(item.split()).rstrip('.')
                    if item_clean and item_clean.lower() not in ('none', 'любая'):
                        # Капитализируем первую букву, сохраняя регистр остальной части (например, для аббревиатур типа АТС, IVR)
                        item_clean = item_clean[0].upper() + item_clean[1:]
                        items.append(item_clean)
                return "|".join(items)

            industries = process_filter('industries')
            products = process_filter('products')
            divisions = process_filter('divisions')

            cases.append({
                'title': title,
                'description': description,
                'industries': industries,
                'products': products,
                'divisions': divisions,
                'url': url
            })

        print(f"Успешно обработано {len(cases)} кейсов из Excel.")

        # Убедимся, что папка назначения существует
        os.makedirs(os.path.dirname(csv_path), exist_ok=True)

        # Запись в CSV
        # Используем utf-8-sig (с BOM) для корректного открытия в Excel на Windows
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            # Записываем заголовок
            writer.writerow(['Название доработки', 'Описание', 'Отрасли', 'Продукты', 'Подразделения', 'Ссылка'])
            # Записываем строки
            for c in cases:
                writer.writerow([
                    c['title'],
                    c['description'],
                    c['industries'],
                    c['products'],
                    c['divisions'],
                    c['url']
                ])

        print(f"Файл {csv_path} успешно обновлен.")

    except Exception as e:
        print(f"Произошла ошибка при обработке: {e}")
        sys.exit(1)

if __name__ == '__main__':
    main()
